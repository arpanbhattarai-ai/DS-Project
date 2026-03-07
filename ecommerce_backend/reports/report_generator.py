"""
reports/report_generator.py  -  Generate PDF and Excel BI reports.
"""
from pathlib import Path
from logger import get_logger

from analysis.profitability import compute_profitability
from analysis.discounts import compute_discounts
from analysis.inventory import compute_inventory
from analysis.products import compute_products
from analysis.expenses import compute_expenses
from analysis.monthly_growth import compute_monthly_growth
from analysis.breakeven import compute_breakeven
from analysis.cashflow import compute_cashflow
from analysis.sales_trend import compute_sales_trend
from analysis.demand_vs_stock import compute_demand_vs_stock
from analysis.cost_efficiency import compute_cost_efficiency

logger = get_logger(__name__)
EXPORT_DIR = Path("data/exports")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _all(store: dict) -> dict:
    context = {
        "inv": store["inv"],
        "sales": store["sales"],
        "purch": store["purch"],
        "exp": store["exp"],
        "analysis": {},
    }

    return {
        "profitability": compute_profitability(store),
        "discounts": compute_discounts(store),
        "inventory": compute_inventory(store),
        "products": compute_products(store),
        "expenses": compute_expenses(store),
        "monthly_growth": compute_monthly_growth(store),
        "breakeven": compute_breakeven(store),
        "cashflow": compute_cashflow(store),
        "sales_trend": compute_sales_trend(context),
        "demand_vs_stock": compute_demand_vs_stock(context),
        "cost_efficiency": compute_cost_efficiency(context),
    }


def _chart_dir() -> Path:
    p = EXPORT_DIR / "_charts"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _save_plot(path: Path, title: str, x, ys: list[tuple[str, list]], kind: str = "line") -> None:
    import matplotlib.pyplot as plt

    plt.figure(figsize=(9, 4.5))
    if kind == "bar":
        if ys:
            label, vals = ys[0]
            plt.bar(x, vals, color="#2563eb", alpha=0.9)
            plt.ylabel(label)
    else:
        for label, vals in ys:
            plt.plot(x, vals, marker="o", linewidth=2, label=label)
        if len(ys) > 1:
            plt.legend()

    plt.title(title, fontsize=12, weight="bold")
    plt.grid(axis="y", linestyle="--", alpha=0.4)
    plt.xticks(rotation=30)
    plt.tight_layout()
    plt.savefig(path, dpi=150)
    plt.close()


def _pdf_section(pdf, title: str) -> None:
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_fill_color(37, 99, 235)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, "  " + title, ln=True, fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 11)
    pdf.ln(2)


def _pdf_row(pdf, label: str, value: str) -> None:
    pdf.cell(95, 7, label)
    pdf.cell(0, 7, str(value), ln=True)


def _render_pdf_visuals(pdf, data: dict) -> None:
    chart_dir = _chart_dir()

    mg = data["monthly_growth"]["monthly"]
    if mg:
        path = chart_dir / "monthly_growth.png"
        _save_plot(
            path,
            "Monthly Revenue and Profit",
            [r["MonthName"] for r in mg],
            [
                ("Revenue (NPR)", [r["MonthRevenue"] for r in mg]),
                ("Profit (NPR)", [r["MonthProfit"] for r in mg]),
            ],
        )
        pdf.image(str(path), w=185)
        pdf.ln(2)

    st = data["sales_trend"]["monthly_sales_revenue"]
    if st:
        path = chart_dir / "sales_trend.png"
        _save_plot(
            path,
            "Monthly Sales Revenue Trend",
            [r["MonthName"] for r in st],
            [("Sales Revenue (NPR)", [r["revenue_npr"] for r in st])],
        )
        pdf.image(str(path), w=185)
        pdf.ln(2)

    cf = data["cashflow"]["monthly_cashflow"]
    if cf:
        path = chart_dir / "cashflow.png"
        _save_plot(
            path,
            "Monthly Cash In, Cash Out, Net Cash",
            [str(r["MonthNum"]) for r in cf],
            [
                ("Cash In", [r["CashIn"] for r in cf]),
                ("Cash Out", [r["CashOut"] for r in cf]),
                ("Net Cash", [r["NetCash"] for r in cf]),
            ],
        )
        pdf.image(str(path), w=185)
        pdf.ln(2)

    ce = data["cost_efficiency"]["monthly_cost_efficiency"]
    if ce:
        path = chart_dir / "cost_efficiency.png"
        _save_plot(
            path,
            "Expense vs Purchase Cost Trend",
            [r["MonthName"] for r in ce],
            [
                ("Purchase Cost", [r["purchase_cost_npr"] for r in ce]),
                ("Operating Expense", [r["operating_expense_npr"] for r in ce]),
            ],
        )
        pdf.image(str(path), w=185)
        pdf.ln(2)

    cat = data["products"]["revenue_by_category"][:8]
    if cat:
        path = chart_dir / "category_revenue.png"
        _save_plot(
            path,
            "Revenue by Category",
            [r["Category"] for r in cat],
            [("Revenue (NPR)", [r["category_revenue_npr"] for r in cat])],
            kind="bar",
        )
        pdf.image(str(path), w=185)
        pdf.ln(2)


def generate_pdf(store: dict) -> str:
    from fpdf import FPDF

    data = _all(store)
    p = data["profitability"]
    d = data["discounts"]
    inv = data["inventory"]
    exp = data["expenses"]
    be = data["breakeven"]
    dvs = data["demand_vs_stock"]
    ce = data["cost_efficiency"]

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 12, "Nepal E-Commerce BI Report 2025", ln=True, align="C")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, "Visual Business Intelligence Summary", ln=True, align="C")
    pdf.ln(3)

    _pdf_section(pdf, "Profitability")
    _pdf_row(pdf, "Gross Revenue", f"NPR {p['gross_revenue_npr']:,.0f}")
    _pdf_row(pdf, "Net Revenue", f"NPR {p['net_revenue_npr']:,.0f}")
    _pdf_row(pdf, "Net Profit", f"NPR {p['net_profit_npr']:,.0f}")
    _pdf_row(pdf, "Net Profit Margin", f"{p['net_profit_margin_pct']}%")

    _pdf_section(pdf, "Discounts")
    _pdf_row(pdf, "Total Discount", f"NPR {d['total_discount_npr']:,.0f}")
    _pdf_row(pdf, "Discount % of Revenue", f"{d['discount_pct_of_revenue']}%")
    _pdf_row(pdf, "Discounted Txn %", f"{d['discounted_txn_pct']}%")

    _pdf_section(pdf, "Inventory")
    _pdf_row(pdf, "Inventory Turnover", f"{inv['inventory_turnover']}x")
    _pdf_row(pdf, "Days Inventory Outstanding", f"{inv['days_inventory_outstanding']} days")
    _pdf_row(pdf, "Items Below Reorder", str(inv["items_below_reorder_level"]))

    _pdf_section(pdf, "Demand vs Stock")
    _pdf_row(pdf, "Items Tracked", dvs["summary"]["items_tracked"])
    _pdf_row(pdf, "Overstock Items", dvs["summary"]["overstock_items"])
    _pdf_row(pdf, "Stockout Risk Items", dvs["summary"]["stockout_risk_items"])

    _pdf_section(pdf, "Expenses")
    _pdf_row(pdf, "Total OpEx", f"NPR {exp['total_opex_npr']:,.0f}")
    _pdf_row(pdf, "OpEx % of Revenue", f"{exp['opex_pct_of_revenue']}%")

    _pdf_section(pdf, "Break-Even")
    _pdf_row(pdf, "Overall Break-Even Units", f"{be['overall_breakeven_units']:,.0f}")
    _pdf_row(pdf, "Margin of Safety", f"{be['margin_of_safety_pct']}%")

    _pdf_section(pdf, "Cost Efficiency")
    _pdf_row(pdf, "Total Purchase Cost", f"NPR {ce['summary']['total_purchase_cost_npr']:,.0f}")
    _pdf_row(pdf, "Total Operating Expense", f"NPR {ce['summary']['total_operating_expense_npr']:,.0f}")

    pdf.add_page()
    _pdf_section(pdf, "Visual Trends")
    _render_pdf_visuals(pdf, data)

    out = str(EXPORT_DIR / "BI_Report.pdf")
    pdf.output(out)
    logger.info(f"PDF saved: {out}")
    return out


def _add_excel_charts(workbook) -> None:
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference

    dash = workbook.create_sheet("Dashboard Charts")
    dash["A1"] = "BI Visual Dashboard"

    # Profitability bar chart
    ws = workbook["Profitability"]
    chart = BarChart()
    chart.title = "Profitability Snapshot"
    chart.y_axis.title = "NPR"
    data = Reference(ws, min_col=2, min_row=2, max_row=10)
    cats = Reference(ws, min_col=1, min_row=2, max_row=10)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    dash.add_chart(chart, "A3")

    # Monthly growth line chart
    ws = workbook["Monthly Growth"]
    chart = LineChart()
    chart.title = "Monthly Revenue vs Profit"
    chart.y_axis.title = "NPR"
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    dash.add_chart(chart, "N3")

    # Expenses pie chart
    ws = workbook["Expenses"]
    chart = PieChart()
    chart.title = "Expense Breakdown"
    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 10
    dash.add_chart(chart, "A20")

    # Cash flow line chart
    ws = workbook["Cash Flow"]
    chart = LineChart()
    chart.title = "Monthly Cashflow"
    data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    dash.add_chart(chart, "N20")

    # Sales trend chart
    ws = workbook["Sales Trend"]
    chart = LineChart()
    chart.title = "Sales Revenue Trend"
    data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    dash.add_chart(chart, "A37")

    # Cost efficiency chart
    ws = workbook["Cost Efficiency"]
    chart = LineChart()
    chart.title = "Expense vs Purchase Cost"
    data = Reference(ws, min_col=3, max_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    dash.add_chart(chart, "N37")


def generate_excel(store: dict) -> str:
    import pandas as pd

    data = _all(store)
    out = str(EXPORT_DIR / "BI_Report.xlsx")

    profitability_df = pd.DataFrame(
        [
            ("gross_revenue_npr", data["profitability"]["gross_revenue_npr"]),
            ("total_discount_npr", data["profitability"]["total_discount_npr"]),
            ("net_revenue_npr", data["profitability"]["net_revenue_npr"]),
            ("cogs_npr", data["profitability"]["cogs_npr"]),
            ("gross_profit_npr", data["profitability"]["gross_profit_npr"]),
            ("gross_profit_margin_pct", data["profitability"]["gross_profit_margin_pct"]),
            ("total_opex_npr", data["profitability"]["total_opex_npr"]),
            ("net_profit_npr", data["profitability"]["net_profit_npr"]),
            ("net_profit_margin_pct", data["profitability"]["net_profit_margin_pct"]),
        ],
        columns=["Metric", "Value"],
    )

    discounts_summary_df = pd.DataFrame(
        [{k: v for k, v in data["discounts"].items() if k != "monthly_discount"}]
    )
    discounts_monthly_df = pd.DataFrame(data["discounts"]["monthly_discount"])

    inventory_summary_df = pd.DataFrame(
        [{k: v for k, v in data["inventory"].items() if k != "below_reorder_items"}]
    )

    expenses_summary_df = pd.DataFrame(
        [{k: v for k, v in data["expenses"].items() if k not in {"expense_breakdown", "monthly_expenses"}}]
    )

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        profitability_df.to_excel(writer, sheet_name="Profitability", index=False)
        discounts_summary_df.to_excel(writer, sheet_name="Discounts Summary", index=False)
        discounts_monthly_df.to_excel(writer, sheet_name="Discounts Monthly", index=False)
        inventory_summary_df.to_excel(writer, sheet_name="Inventory Summary", index=False)

        pd.DataFrame(data["monthly_growth"]["monthly"]).to_excel(
            writer, sheet_name="Monthly Growth", index=False
        )
        pd.DataFrame(data["products"]["top_10_products_by_revenue"]).to_excel(
            writer, sheet_name="Top Products", index=False
        )
        pd.DataFrame(data["products"]["revenue_by_category"]).to_excel(
            writer, sheet_name="Category Revenue", index=False
        )

        pd.DataFrame([data["expenses"]["expense_breakdown"]]).T.reset_index().rename(
            columns={"index": "Category", 0: "NPR"}
        ).to_excel(writer, sheet_name="Expenses", index=False)
        expenses_summary_df.to_excel(writer, sheet_name="Expenses Summary", index=False)
        pd.DataFrame(data["expenses"]["monthly_expenses"]).to_excel(
            writer, sheet_name="Monthly Expenses", index=False
        )

        pd.DataFrame(data["cashflow"]["monthly_cashflow"]).to_excel(
            writer, sheet_name="Cash Flow", index=False
        )
        pd.DataFrame([{k: v for k, v in data["cashflow"].items() if k != "monthly_cashflow"}]).to_excel(
            writer, sheet_name="Cash Flow Summary", index=False
        )

        pd.DataFrame(data["breakeven"]["top_20_easiest_breakeven"]).to_excel(
            writer, sheet_name="Break-Even", index=False
        )
        pd.DataFrame([{k: v for k, v in data["breakeven"].items() if k != "top_20_easiest_breakeven"}]).to_excel(
            writer, sheet_name="Break-Even Summary", index=False
        )

        pd.DataFrame(data["inventory"]["below_reorder_items"]).to_excel(
            writer, sheet_name="Reorder Alerts", index=False
        )

        pd.DataFrame(data["sales_trend"]["monthly_sales_revenue"]).to_excel(
            writer, sheet_name="Sales Trend", index=False
        )
        pd.DataFrame(data["sales_trend"]["monthly_by_payment_mode"]).to_excel(
            writer, sheet_name="Sales by Payment", index=False
        )
        pd.DataFrame(data["sales_trend"]["monthly_by_delivery_type"]).to_excel(
            writer, sheet_name="Sales by Delivery", index=False
        )

        pd.DataFrame(data["demand_vs_stock"]["detail"]).to_excel(
            writer, sheet_name="Demand vs Stock", index=False
        )
        pd.DataFrame(data["demand_vs_stock"]["overstock_items"]).to_excel(
            writer, sheet_name="Overstock Items", index=False
        )
        pd.DataFrame(data["demand_vs_stock"]["stockout_risk_items"]).to_excel(
            writer, sheet_name="Stockout Risk", index=False
        )
        pd.DataFrame([data["demand_vs_stock"]["summary"]]).to_excel(
            writer, sheet_name="Demand Stock Summary", index=False
        )

        pd.DataFrame(data["cost_efficiency"]["monthly_cost_efficiency"]).to_excel(
            writer, sheet_name="Cost Efficiency", index=False
        )
        pd.DataFrame([data["cost_efficiency"]["summary"]]).to_excel(
            writer, sheet_name="Cost Efficiency Summary", index=False
        )

        _add_excel_charts(writer.book)

    logger.info(f"Excel saved: {out}")
    return out
